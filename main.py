import os
import io
import re
import shutil
import uvicorn
from contextlib import asynccontextmanager
from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Query, Depends
from fastapi.responses import HTMLResponse, FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from typing import Optional, List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from database import (
    init_db, get_all_teachers, get_teacher_by_id,
    create_teacher, update_teacher, delete_teacher, get_stats,
    get_db_connection, DB_PATH
)
from importer import scan_directory, parse_docx_file, parse_xlsx_file, merge_and_save_data, UPLOAD_DIR, PHOTO_DIR
from models import TeacherCreate, TeacherUpdate, BackupResponse
from auth import verify_admin_key

# Maximum upload file size: 10 MB
MAX_UPLOAD_SIZE = 10 * 1024 * 1024

# Allowed image MIME types
ALLOWED_IMAGE_TYPES = {'image/jpeg', 'image/png', 'image/gif', 'image/webp'}

def sanitize_filename(filename: str) -> str:
    """Remove path traversal and dangerous characters from filename"""
    # Strip directory components
    filename = os.path.basename(filename)
    # Keep only safe characters: Thai, alphanumeric, dots, hyphens, underscores
    filename = re.sub(r'[^\w.\-฀-๿]', '_', filename)
    # Prevent double dots (relative path traversal)
    filename = re.sub(r'\.\.+', '.', filename)
    # Limit length
    if len(filename) > 200:
        name, ext = os.path.splitext(filename)
        filename = name[:190] + ext
    return filename or 'unnamed_file' 

@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    stats = get_stats()
    if stats["teacher_count"] == 0:
        base_dir = os.path.dirname(__file__)
        scan_directory(base_dir)
    yield

app = FastAPI(title="ระบบบริหารจัดการครู อาจารย์ ครูฝึก", version="1.0.0", lifespan=lifespan)

# Enable CORS
# In production, set ALLOWED_ORIGINS env var (comma-separated) for your domain
_allowed_origins_env = os.environ.get("ALLOWED_ORIGINS", "")
if _allowed_origins_env:
    _allowed_origins = [o.strip() for o in _allowed_origins_env.split(",") if o.strip()]
else:
    # Default: allow localhost for development
    _allowed_origins = ["http://localhost:8000", "http://127.0.0.1:8000"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=_allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Ensure upload folders exist
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(PHOTO_DIR, exist_ok=True)
static_dir = os.path.join(os.path.dirname(__file__), "static")
os.makedirs(static_dir, exist_ok=True)

# Mount static files
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")
app.mount("/static", StaticFiles(directory=static_dir), name="static")

@app.get("/")
def get_index():
    index_path = os.path.join(static_dir, "index.html")
    if os.path.exists(index_path):
        return FileResponse(index_path, media_type="text/html; charset=utf-8")
    return HTMLResponse("<h1>ระบบบริหารจัดการครูอาจารย์กำลังเริ่มต้น...</h1>")

@app.get("/api/stats")
def api_get_stats():
    return get_stats()

@app.get("/api/teachers")
def api_get_teachers(
    q: Optional[str] = Query(None),
    course: Optional[str] = Query(None)
):
    return get_all_teachers(query=q, course_filter=course)

@app.get("/api/teachers/{teacher_id}")
def api_get_teacher(teacher_id: int):
    teacher = get_teacher_by_id(teacher_id)
    if not teacher:
        raise HTTPException(status_code=404, detail="Teacher not found")
    return teacher

@app.post("/api/teachers")
def api_create_teacher(data: TeacherCreate, _auth: str = Depends(verify_admin_key)):
    teacher_id = create_teacher(data.model_dump())
    return {"status": "success", "id": teacher_id}

@app.put("/api/teachers/{teacher_id}")
def api_update_teacher(teacher_id: int, data: TeacherUpdate, _auth: str = Depends(verify_admin_key)):
    success = update_teacher(teacher_id, data.model_dump())
    if not success:
        raise HTTPException(status_code=404, detail="Teacher not found")
    return {"status": "success", "id": teacher_id}

@app.delete("/api/teachers/{teacher_id}")
def api_delete_teacher(teacher_id: int, _auth: str = Depends(verify_admin_key)):
    success = delete_teacher(teacher_id)
    if not success:
        raise HTTPException(status_code=404, detail="Teacher not found")
    return {"status": "success"}

@app.post("/api/scan")
def api_scan_folder(_auth: str = Depends(verify_admin_key)):
    base_dir = os.path.dirname(__file__)
    result = scan_directory(base_dir)
    return {"status": "success", "result": result}

@app.post("/api/upload-file")
async def api_upload_file(file: UploadFile = File(...), _auth: str = Depends(verify_admin_key)):
    filename = sanitize_filename(file.filename or "unnamed")
    
    # Validate file extension
    if not (filename.lower().endswith(".docx") or filename.lower().endswith(".xlsx")):
        raise HTTPException(status_code=400, detail="รองรับเฉพาะไฟล์ .docx หรือ .xlsx เท่านั้น")
    
    # Check file size
    content_bytes = await file.read()
    if len(content_bytes) > MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=413, detail="ไฟล์มีขนาดใหญ่เกิน 10 MB")
    
    temp_path = os.path.join(UPLOAD_DIR, filename)
    with open(temp_path, "wb") as buffer:
        buffer.write(content_bytes)
        
    records = []
    if filename.lower().endswith(".docx"):
        records = parse_docx_file(temp_path)
    elif filename.lower().endswith(".xlsx"):
        records = parse_xlsx_file(temp_path)
        
    result = merge_and_save_data(records)
    return {"status": "success", "result": result, "filename": filename}

@app.post("/api/upload-photo/{teacher_id}")
async def api_upload_photo(teacher_id: int, photo: UploadFile = File(...), _auth: str = Depends(verify_admin_key)):
    teacher = get_teacher_by_id(teacher_id)
    if not teacher:
        raise HTTPException(status_code=404, detail="Teacher not found")
    
    # Validate image type
    if photo.content_type not in ALLOWED_IMAGE_TYPES:
        raise HTTPException(status_code=400, detail="รองรับเฉพาะไฟล์รูปภาพ (JPEG, PNG, GIF, WebP) เท่านั้น")
    
    # Check file size
    photo_bytes = await photo.read()
    if len(photo_bytes) > MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=413, detail="ไฟล์รูปมีขนาดใหญ่เกิน 10 MB")
    
    # Sanitize and save
    safe_name = sanitize_filename(photo.filename or "photo")
    ext = os.path.splitext(safe_name)[1] or ".jpg"
    filename = f"teacher_{teacher_id}_{safe_name}"
    filepath = os.path.join(PHOTO_DIR, filename)
    
    with open(filepath, "wb") as buffer:
        buffer.write(photo_bytes)
        
    photo_url = f"/uploads/photos/{filename}"
    teacher["photo_url"] = photo_url
    update_teacher(teacher_id, teacher)
    
    return {"status": "success", "photo_url": photo_url}

@app.get("/api/export-excel")
def api_export_excel(course: Optional[str] = Query(None)):
    teachers = get_all_teachers(course_filter=course)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "บัญชีรายชื่อครูอาจารย์"
    
    # Title
    course_title = f"หลักสูตร {course}" if course and course != 'all' else "ทุกหลักสูตร"
    ws.merge_cells("A1:I1")
    ws["A1"] = f"บัญชีรายชื่อครู อาจารย์ ครูฝึก ({course_title})"
    ws["A1"].font = Font(name="TH SarabunPSK", size=18, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ["ลำดับ", "ยศ", "ชื่อ", "สกุล", "ตำแหน่ง/สังกัด", "วุฒิการศึกษา", "หลักสูตรที่สอน", "วิชาที่สอน", "หมายเหตุ"]
    ws.append([]) # Row 2 empty
    ws.append(headers) # Row 3
    
    header_font = Font(name="TH SarabunPSK", size=14, bold=True)
    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    # Batch fetch teaching assignments for all teachers (eliminates N+1 query)
    from database import get_db_connection
    teacher_ids = [t["id"] for t in teachers]
    assignments_map = {}
    if teacher_ids:
        conn = get_db_connection()
        cursor = conn.cursor()
        placeholders = ','.join('?' * len(teacher_ids))
        cursor.execute(
            f"SELECT teacher_id, subject_name, subject_category FROM teaching_assignments WHERE teacher_id IN ({placeholders})",
            teacher_ids
        )
        for row in cursor.fetchall():
            tid = row["teacher_id"]
            if tid not in assignments_map:
                assignments_map[tid] = []
            assignments_map[tid].append(dict(row))
        conn.close()
    
        row_font = Font(name="TH SarabunPSK", size=13)
    for idx, t in enumerate(teachers, 1):
        t_id = t["id"]
        teaching_assignments = assignments_map.get(t_id, [])
        
        # Format courses
        courses_str = "\n".join([f"{i+1}. {c}" for i, c in enumerate(t["courses"])]) if t["courses"] else "-"
        
        # Format subjects
        subjects_list = []
        if teaching_assignments:
            academic_subj = [a["subject_name"] for a in teaching_assignments if a.get("subject_category") == "ภาควิชาการ" and a.get("subject_name")]
            drill_subj = [a["subject_name"] for a in teaching_assignments if a.get("subject_category") == "ภาคการฝึก" and a.get("subject_name")]
            other_subj = [a["subject_name"] for a in teaching_assignments if not a.get("subject_category") and a.get("subject_name")]
            
            if academic_subj:
                subjects_list.append("ภาควิชาการ")
                for i, s in enumerate(list(dict.fromkeys(academic_subj))):
                    subjects_list.append(f"{i+1}. {s}")
            if drill_subj:
                subjects_list.append("ภาคการฝึก")
                for i, s in enumerate(list(dict.fromkeys(drill_subj))):
                    subjects_list.append(f"{i+1}. {s}")
            if other_subj and not academic_subj and not drill_subj:
                for i, s in enumerate(list(dict.fromkeys(other_subj))):
                    subjects_list.append(f"{i+1}. {s}")
                    
        subjects_str = "\n".join(subjects_list) if subjects_list else "-"
        
        pos_full = f"{t.get('position', '')} {t.get('affiliation', '')}".strip()
        
        row_data = [
            idx,
            t.get("prefix_rank", ""),
            t.get("first_name", ""),
            t.get("last_name", ""),
            pos_full,
            t.get("education_summary", ""),
            courses_str,
            subjects_str,
            t.get("notes", "")
        ]
        ws.append(row_data)
        
        current_row = ws.max_row
        for c_idx in range(1, len(row_data) + 1):
            c_cell = ws.cell(row=current_row, column=c_idx)
            c_cell.font = row_font
            c_cell.border = thin_border
            c_cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="center" if c_idx in [1, 2] else "left")
            
    # Adjust column widths
    col_widths = [8, 12, 16, 18, 30, 25, 20, 35, 15]
    for i, w in enumerate(col_widths, 1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = w
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=teacher_roster.xlsx"}
    )

# =========================================================================
# DEBUG: LIST AVAILABLE FILES
# =========================================================================

@app.get("/api/debug/files")
def api_debug_files():
    """List all .docx and .xlsx files available for scanning."""
    import glob as _glob
    base_dir = os.path.dirname(__file__)
    files = []
    for pattern in ['**/*.docx', '**/*.xlsx']:
        for f in _glob.glob(os.path.join(base_dir, pattern), recursive=True):
            if '__pycache__' in f or '.freebuff' in f or f.startswith('~'):
                continue
            files.append({
                'path': os.path.relpath(f, base_dir),
                'size': os.path.getsize(f)
            })
    return {'base_dir': base_dir, 'files': files, 'count': len(files)}


# =========================================================================
# RESET DATABASE & RE-SCAN
# =========================================================================

@app.post("/api/reset-and-scan")
def api_reset_and_scan(_auth: str = Depends(verify_admin_key)):
    """Delete database and re-scan all files with updated importer logic."""
    import shutil as _shutil
    
    # Delete existing database
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
        for suffix in ['-wal', '-shm', '-journal']:
            fpath = DB_PATH + suffix
            if os.path.exists(fpath):
                os.remove(fpath)
    
    # Re-initialize DB
    init_db()
    
    # Re-scan all files
    base_dir = os.path.dirname(__file__)
    try:
        result = scan_directory(base_dir)
        return {
            "status": "success",
            "message": "ล้างข้อมูลและสแกนใหม่เรียบร้อย",
            "result": result
        }
    except Exception as e:
        return {
            "status": "error",
            "message": f"เกิดข้อผิดพลาด: {str(e)}",
            "result": {"total_processed": 0, "new_count": 0, "merged_count": 0, "errors": [str(e)]}
        }


# =========================================================================
# BACKUP & RESTORE
# =========================================================================

@app.get("/api/backup")
def api_backup(_auth: str = Depends(verify_admin_key)):
    """Create a timestamped backup of the database."""
    import shutil as _shutil
    from datetime import datetime
    
    if not os.path.exists(DB_PATH):
        raise HTTPException(status_code=404, detail="Database file not found")
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_filename = f"teachers_backup_{timestamp}.db"
    backup_dir = os.path.join(os.path.dirname(__file__), "backups")
    os.makedirs(backup_dir, exist_ok=True)
    backup_path = os.path.join(backup_dir, backup_filename)
    
    _shutil.copy2(DB_PATH, backup_path)
    file_size = os.path.getsize(backup_path)
    
    return BackupResponse(
        status="success",
        filename=backup_filename,
        size=file_size
    )


@app.get("/api/backups")
def api_list_backups(_auth: str = Depends(verify_admin_key)):
    """List available backup files."""
    backup_dir = os.path.join(os.path.dirname(__file__), "backups")
    if not os.path.exists(backup_dir):
        return {"backups": []}
    
    files = []
    for f in sorted(os.listdir(backup_dir), reverse=True):
        if f.startswith("teachers_backup_") and f.endswith(".db"):
            fpath = os.path.join(backup_dir, f)
            files.append({
                "filename": f,
                "size": os.path.getsize(fpath),
            })
    return {"backups": files}


@app.post("/api/restore")
def api_restore(filename: str = Query(...), _auth: str = Depends(verify_admin_key)):
    """Restore database from a backup file."""
    import shutil as _shutil
    
    backup_dir = os.path.join(os.path.dirname(__file__), "backups")
    backup_path = os.path.join(backup_dir, filename)
    
    # Security: only allow restoring from backup files in the backups dir
    if not filename.startswith("teachers_backup_") or not filename.endswith(".db"):
        raise HTTPException(status_code=400, detail="Invalid backup filename")
    if not os.path.exists(backup_path):
        raise HTTPException(status_code=404, detail="Backup file not found")
    if not os.path.abspath(backup_path).startswith(os.path.abspath(backup_dir)):
        raise HTTPException(status_code=400, detail="Path traversal not allowed")
    
    _shutil.copy2(backup_path, DB_PATH)
    return {"status": "success", "message": f"Database restored from {filename}"}


@app.get("/api/export-db")
def api_export_db(_auth: str = Depends(verify_admin_key)):
    """Download the entire database file."""
    if not os.path.exists(DB_PATH):
        raise HTTPException(status_code=404, detail="Database file not found")
    
    return FileResponse(
        DB_PATH,
        media_type="application/octet-stream",
        filename="teachers_database.db"
    )


if __name__ == "__main__":
    # Use 0.0.0.0 to accept connections from any host
    # Set HOST and PORT env vars to customize
    host = os.environ.get("HOST", "0.0.0.0")
    port = int(os.environ.get("PORT", "8000"))
    uvicorn.run("main:app", host=host, port=port, reload=True)
