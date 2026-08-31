import os
import re
import io
import shutil
import sqlite3
import docx
import openpyxl
from xml.etree import ElementTree as ET
from typing import List, Dict, Any, Tuple
from database import get_db_connection, init_db

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads")
PHOTO_DIR = os.path.join(UPLOAD_DIR, "photos")
os.makedirs(PHOTO_DIR, exist_ok=True)

def clean_text(text: str) -> str:
    if not text:
        return ""
    text = re.sub(r'[\u200b\u200c\u200d\uFEFF]', '', str(text))
    return text.strip()

def normalize_name(prefix: str, first: str, last: str) -> Tuple[str, str, str]:
    prefix = clean_text(prefix).replace('หณิง', 'หญิง')
    first = clean_text(first)
    last = clean_text(last)
    
    # Check if prefix leaked into first name
    for p in ['พ.ต.อ.', 'พ.ต.ท.', 'พ.ต.ต.', 'ร.ต.อ.หญิง', 'ร.ต.อ.', 'ร.ต.ท.', 'ร.ต.ต.', 'ด.ต.', 'จ.ส.ต.หญิง', 'จ.ส.ต.', 'ส.ต.อ.', 'ส.ต.ท.', 'ส.ต.ต.', 'นาย', 'นาง', 'นางสาว']:
        if first.startswith(p) and not prefix:
            prefix = p
            first = first[len(p):].strip()
            
    return prefix, first, last

def extract_images_from_docx_element(elem, doc, prefix_name: str) -> str:
    """Extract image if referenced in XML element and save to disk"""
    try:
        # Search for blip elements in XML
        xml_str = elem._element.xml
        blip_matches = re.findall(r'r:embed="([^"]+)"', xml_str)
        for r_id in blip_matches:
            if r_id in doc.part.rels:
                rel = doc.part.rels[r_id]
                if "image" in rel.target_ref:
                    image_part = rel.target_part
                    ext = image_part.content_type.split('/')[-1]
                    if ext == 'jpeg': ext = 'jpg'
                    filename = f"photo_{clean_text(prefix_name)}_{r_id}.{ext}"
                    filepath = os.path.join(PHOTO_DIR, filename)
                    with open(filepath, "wb") as f:
                        f.write(image_part.blob)
                    return f"/uploads/photos/{filename}"
    except Exception as e:
        print(f"Error extracting image: {e}")
    return ""

def parse_docx_file(filepath: str) -> List[Dict[str, Any]]:
    doc = docx.Document(filepath)
    filename_base = os.path.splitext(os.path.basename(filepath))[0]
    
    # Split docx by 'ประวัติ ครู อาจารย์ ครูฝึก'
    sections = []
    current_sec = {'elements': [], 'paragraphs': [], 'tables': []}
    
    for child in doc._body._body:
        tag = child.tag.split('}')[-1]
        if tag == 'p':
            p_elem = docx.text.paragraph.Paragraph(child, doc)
            txt = p_elem.text.strip()
            # Split on section header (check multiple patterns)
            if ('ประวัติ ครู อาจารย์ ครูฝึก' in txt or 
                'ประวัติครูอาจารย์' in txt or
                (txt.strip() == '' and current_sec.get('_seen_header', False))):
                if current_sec['paragraphs'] or current_sec['tables']:
                    sections.append(current_sec)
                current_sec = {'elements': [p_elem], 'paragraphs': [p_elem], 'tables': [], '_seen_header': True}
            else:
                current_sec['elements'].append(p_elem)
                if txt:
                    current_sec['paragraphs'].append(p_elem)
        elif tag == 'tbl':
            tbl_elem = docx.table.Table(child, doc)
            current_sec['elements'].append(tbl_elem)
            current_sec['tables'].append(tbl_elem)
            
    if current_sec['paragraphs'] or current_sec['tables']:
        sections.append(current_sec)
        
    teachers = []
    for sec_idx, sec in enumerate(sections):
        p_texts = [p.text for p in sec['paragraphs']]
        full_text = '\n'.join(p_texts)
        
        # Check if contains teacher profile
        if not ('ชื่อ' in full_text and 'สกุล' in full_text):
            continue
            
        # Parse Personal Info
        # Prefix, First, Last, Age
        prefix, first_name, last_name, age = '', '', '', ''
        m_name = re.search(r'(?:ยศ|คำนำหน้า)\s*([^\s]+(?:\s+หญิง|\s+หณิง)?)\s*ชื่อ\s*([^\s]+)\s*สกุล\s*([^\s]+)(?:\s*อายุ\s*([^\s]*?)\s*ปี)?', full_text)
        if m_name:
            prefix = m_name.group(1).replace('หณิง', 'หญิง').strip()
            first_name = m_name.group(2).strip()
            last_name = m_name.group(3).strip()
            if m_name.group(4):
                age = m_name.group(4).strip()
                if age in ['-', '–', '.']: age = ''
        else:
            # Fallback
            m_f = re.search(r'ชื่อ\s*([^\s\t\n]+)', full_text)
            m_l = re.search(r'สกุล\s*([^\s\t\n]+)', full_text)
            m_p = re.search(r'(?:ยศ|คำนำหน้า)\s*([^\s\t\n]+(?:\s+หญิง)?)', full_text)
            if m_f: first_name = m_f.group(1).strip()
            if m_l: last_name = m_l.group(1).strip()
            if m_p: prefix = m_p.group(1).strip()
            
        prefix, first_name, last_name = normalize_name(prefix, first_name, last_name)
        if not first_name or not last_name or first_name == 'สกุล':
            continue
            
        # Extract Photo (from paragraphs AND table cells)
        photo_url = ""
        # Search paragraphs first
        for p in sec['paragraphs']:
            img_url = extract_images_from_docx_element(p, doc, f"{first_name}_{last_name}_{sec_idx}")
            if img_url:
                photo_url = img_url
                break
        # If no photo in paragraphs, search table cells
        if not photo_url:
            for table in sec.get('tables', []):
                for row in table.rows:
                    for cell in row.cells:
                        for p in cell.paragraphs:
                            img_url = extract_images_from_docx_element(p, doc, f"{first_name}_{last_name}_{sec_idx}")
                            if img_url:
                                photo_url = img_url
                                break
                        if photo_url:
                            break
                    if photo_url:
                        break
                if photo_url:
                    break
                
        # Position & Affiliation
        position, affiliation = '', ''
        m_pos = re.search(r'ตำแหน่ง\s*([^\t\n\rสังกัด]+?)(?:\s*สังกัด\s*([^\n\r]+))?(?:\n|$)', full_text)
        if m_pos:
            position = clean_text(m_pos.group(1)).rstrip('.')
            if m_pos.group(2):
                affiliation = clean_text(m_pos.group(2)).rstrip('.')
                
        # Workplace Address
        addr_parts = []
        m_work = re.search(r'สถานที่ทำงาน\s*([^\n\r]+)', full_text)
        if m_work:
            addr_parts.append(re.sub(r'\s+', ' ', clean_text(m_work.group(1))).rstrip('.'))
        m_amp = re.search(r'อำเภอ\s*([^\n\r]+)', full_text)
        if m_amp:
            addr_parts.append(re.sub(r'\s+', ' ', clean_text(m_amp.group(1))).rstrip('.'))
        workplace_address = ' '.join(addr_parts).strip()
        
        # Phone & Email
        phone, email = '', ''
        m_phone = re.search(r'หมายเลขโทรศัพท์\s*([^\t\n\rE\-mail]+?)(?:\s*E-mail\s*([^\n\r]+))?(?:\n|$)', full_text, re.IGNORECASE)
        if m_phone:
            phone = clean_text(m_phone.group(1)).strip('-').strip('.').strip()
            if m_phone.group(2):
                email = clean_text(m_phone.group(2)).strip('-').strip('.').strip()
                
        # Tables Parsing
        educations = []
        trainings = []
        work_histories = []
        teaching_assignments = []
        
        for table in sec['tables']:
            if not table.rows:
                continue
            header_cells = [clean_text(c.text) for c in table.rows[0].cells]
            header_str = ' '.join(header_cells)
            
            # Education table
            if 'วุฒิ' in header_str or 'ระดับ' in header_str or 'สาขา' in header_str:
                for r in table.rows[1:]:
                    cells = [clean_text(c.text) for c in r.cells]
                    if len(cells) >= 4:
                        y, lvl, deg, inst = cells[0], cells[1], cells[2], cells[3]
                        if y == 'พ.ศ. -' or y == '-': y = ''
                        if any([y, lvl, deg, inst]):
                            educations.append({
                                'year': y, 'level': lvl, 'degree_field': deg, 'institution': inst
                            })
                            
            # Training table
            elif ('ฝึกอบรม' in header_str) or ('หลักสูตร' in header_str and ('ฝึกอบรมโดย' in header_str or 'สถานที่ฝึกอบรม' in header_str)):
                for r in table.rows[1:]:
                    cells = [clean_text(c.text) for c in r.cells]
                    if len(cells) >= 3:
                        y, course, org = cells[0], cells[1], cells[2]
                        if any([y, course, org]):
                            trainings.append({
                                'year': y, 'course_name': course, 'organized_by': org
                            })
                            
            # Work history table
            elif 'ว/ด/ป' in header_str or ('ตำแหน่ง' in header_str and len(table.rows[0].cells) == 2):
                for r in table.rows[1:]:
                    cells = [clean_text(c.text) for c in r.cells]
                    if len(cells) >= 2:
                        dt, pos_role = cells[0], cells[1]
                        if any([dt, pos_role]):
                            work_histories.append({
                                'date_period': dt, 'position_role': pos_role
                            })
                            
            # Teaching history table
            elif 'ประวัติการสอน' in header_str or ('หลักสูตร' in header_str and 'วิชา' in header_str):
                for r in table.rows[1:]:
                    cells = [clean_text(c.text) for c in r.cells]
                    if len(cells) >= 2:
                        c_type, subj = cells[0], cells[1]
                        if any([c_type, subj]):
                            # might contain multiple subjects separated by comma or newline
                            subjects = [s.strip() for s in re.split(r'[,;\n]', subj) if s.strip()]
                            if not subjects:
                                teaching_assignments.append({
                                    'course_type': c_type,
                                    'subject_category': '',
                                    'subject_name': ''
                                })
                            else:
                                for s in subjects:
                                    teaching_assignments.append({
                                        'course_type': c_type,
                                        'subject_category': '',
                                        'subject_name': s
                                    })
                                    
        teacher_data = {
            'prefix_rank': prefix,
            'first_name': first_name,
            'last_name': last_name,
            'age': age,
            'position': position,
            'affiliation': affiliation,
            'workplace_address': workplace_address,
            'phone': phone,
            'email': email,
            'photo_url': photo_url,
            'educations': educations,
            'trainings': trainings,
            'work_histories': work_histories,
            'teaching_assignments': teaching_assignments,
            'source_file': filepath
        }
        teachers.append(teacher_data)
        
    return teachers

def parse_xlsx_file(filepath: str) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    # Find header row
    header_row_idx = None
    col_map = {}
    
    for r in range(1, min(ws.max_row + 1, 10)):
        row_vals = [clean_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        row_str = ' '.join(row_vals)
        if 'ชื่อ' in row_str and 'สกุล' in row_str:
            header_row_idx = r
            for c in range(1, ws.max_column + 1):
                val = clean_text(ws.cell(r, c).value)
                if 'ลำดับ' in val: col_map['index'] = c
                elif 'ยศ' in val: col_map['prefix'] = c
                elif 'ชื่อ' in val: col_map['first_name'] = c
                elif 'สกุล' in val: col_map['last_name'] = c
                elif 'ตำแหน่ง' in val: col_map['position'] = c
                elif 'วุฒิ' in val: col_map['education'] = c
                elif 'หลักสูตร' in val: col_map['courses'] = c
                elif 'วิชา' in val: col_map['subjects'] = c
                elif 'หมายเหตุ' in val: col_map['notes'] = c
            break
            
    if not header_row_idx:
        return []
        
    teachers = []
    
    for r in range(header_row_idx + 1, ws.max_row + 1):
        first_name = clean_text(ws.cell(r, col_map.get('first_name', 3)).value) if 'first_name' in col_map else ''
        last_name = clean_text(ws.cell(r, col_map.get('last_name', 4)).value) if 'last_name' in col_map else ''
        
        # Stop or skip if empty or signature block
        if not first_name or 'ตรวจแล้วถูกต้อง' in first_name or 'ผู้ตรวจ' in first_name:
            continue
            
        prefix = clean_text(ws.cell(r, col_map.get('prefix', 2)).value) if 'prefix' in col_map else ''
        pos_raw = clean_text(ws.cell(r, col_map.get('position', 5)).value) if 'position' in col_map else ''
        edu_raw = clean_text(ws.cell(r, col_map.get('education', 6)).value) if 'education' in col_map else ''
        courses_raw = clean_text(ws.cell(r, col_map.get('courses', 7)).value) if 'courses' in col_map else ''
        subj_raw = clean_text(ws.cell(r, col_map.get('subjects', 8)).value) if 'subjects' in col_map else ''
        notes = clean_text(ws.cell(r, col_map.get('notes', 9)).value) if 'notes' in col_map else ''
        
        prefix, first_name, last_name = normalize_name(prefix, first_name, last_name)
        
        # Split position & affiliation
        position = pos_raw
        affiliation = ''
        if 'บก.' in pos_raw or 'กก.' in pos_raw or 'บช.' in pos_raw:
            parts = pos_raw.split('บก.')
            if len(parts) == 2:
                position = parts[0].strip()
                affiliation = ('บก.' + parts[1]).strip()
                
        # Parse education
        educations = []
        if edu_raw:
            lvl = ''
            deg = edu_raw
            if 'ปริญญาตรี' in edu_raw:
                lvl = 'ป.ตรี'
            elif 'ปริญญาโท' in edu_raw:
                lvl = 'ป.โท'
            elif 'ปริญญาเอก' in edu_raw:
                lvl = 'ป.เอก'
            elif 'ม.ปลาย' in edu_raw:
                lvl = 'ม.ปลาย'
            educations.append({
                'year': '',
                'level': lvl,
                'degree_field': deg,
                'institution': ''
            })
            
        # Parse courses
        courses = []
        if courses_raw:
            # Pattern: 1. นรต. \n 2. นสต.
            lines = [re.sub(r'^\d+[\.\)]\s*', '', l).strip() for l in courses_raw.split('\n') if l.strip()]
            courses = [c for c in lines if c]
            
        # Parse subjects & category
        teaching_assignments = []
        subject_category = 'ภาควิชาการ'
        if subj_raw:
            lines = subj_raw.split('\n')
            for l in lines:
                l_clean = l.strip()
                if not l_clean: continue
                if 'ภาควิชาการ' in l_clean:
                    subject_category = 'ภาควิชาการ'
                    continue
                elif 'ภาคการฝึก' in l_clean or 'ภาคปฏิบัติ' in l_clean:
                    subject_category = 'ภาคการฝึก'
                    continue
                s_name = re.sub(r'^\d+[\.\)]\s*', '', l_clean).strip()
                if s_name:
                    if courses:
                        for c in courses:
                            teaching_assignments.append({
                                'course_type': c,
                                'subject_category': subject_category,
                                'subject_name': s_name
                            })
                    else:
                        teaching_assignments.append({
                            'course_type': '',
                            'subject_category': subject_category,
                            'subject_name': s_name
                        })
        elif courses:
            for c in courses:
                teaching_assignments.append({
                    'course_type': c,
                    'subject_category': '',
                    'subject_name': ''
                })
                
        teachers.append({
            'prefix_rank': prefix,
            'first_name': first_name,
            'last_name': last_name,
            'position': position,
            'affiliation': affiliation,
            'notes': notes,
            'educations': educations,
            'trainings': [],
            'work_histories': [],
            'teaching_assignments': teaching_assignments,
            'source_file': filepath
        })
        
    return teachers

def merge_and_save_data(teacher_records: List[Dict[str, Any]]) -> Dict[str, Any]:
    init_db()
    conn = get_db_connection()
    cursor = conn.cursor()
    
    new_count = 0
    merged_count = 0
    
    for rec in teacher_records:
        first = clean_text(rec["first_name"])
        last = clean_text(rec["last_name"])
        if not first or not last:
            continue
            
        # Search existing teacher by first & last name
        cursor.execute("SELECT * FROM teachers WHERE first_name = ? AND last_name = ?", (first, last))
        existing = cursor.fetchone()
        
        if existing:
            # MERGE INTO EXISTING RECORD
            t_id = existing["id"]
            merged_count += 1
            
            # Update fields if existing is empty or new has more info
            prefix = rec.get("prefix_rank") or existing["prefix_rank"]
            age = rec.get("age") or existing["age"]
            pos = rec.get("position") or existing["position"]
            affil = rec.get("affiliation") or existing["affiliation"]
            addr = rec.get("workplace_address") or existing["workplace_address"]
            phone = rec.get("phone") or existing["phone"]
            email = rec.get("email") or existing["email"]
            photo = rec.get("photo_url") or existing["photo_url"]
            notes = rec.get("notes") or existing["notes"]
            
            cursor.execute("""
            UPDATE teachers SET
                prefix_rank = ?, age = ?, position = ?, affiliation = ?,
                workplace_address = ?, phone = ?, email = ?, photo_url = ?, notes = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """, (prefix, age, pos, affil, addr, phone, email, photo, notes, t_id))
            
            # Merge Educations
            cursor.execute("SELECT * FROM educations WHERE teacher_id = ?", (t_id,))
            exist_edus = [dict(e) for e in cursor.fetchall()]
            for edu in rec.get("educations", []):
                deg = clean_text(edu.get("degree_field", ""))
                lvl = clean_text(edu.get("level", ""))
                inst = clean_text(edu.get("institution", ""))
                # Check duplicate
                dup = any(
                    (deg and deg in clean_text(e["degree_field"])) or
                    (inst and inst in clean_text(e["institution"]) and deg and deg in clean_text(e["degree_field"]))
                    for e in exist_edus
                )
                if not dup and any([edu.get("year"), lvl, deg, inst]):
                    cursor.execute("""
                    INSERT INTO educations (teacher_id, year, level, degree_field, institution)
                    VALUES (?, ?, ?, ?, ?)
                    """, (t_id, edu.get("year", ""), lvl, deg, inst))
                    
            # Merge Trainings
            cursor.execute("SELECT * FROM trainings WHERE teacher_id = ?", (t_id,))
            exist_trainings = [dict(tr) for tr in cursor.fetchall()]
            for tr in rec.get("trainings", []):
                c_name = clean_text(tr.get("course_name", ""))
                dup = any(c_name and c_name == clean_text(e["course_name"]) for e in exist_trainings)
                if not dup and c_name:
                    cursor.execute("""
                    INSERT INTO trainings (teacher_id, year, course_name, organized_by)
                    VALUES (?, ?, ?, ?)
                    """, (t_id, tr.get("year", ""), c_name, tr.get("organized_by", "")))
                    
            # Merge Work Histories
            cursor.execute("SELECT * FROM work_histories WHERE teacher_id = ?", (t_id,))
            exist_works = [dict(w) for w in cursor.fetchall()]
            for w in rec.get("work_histories", []):
                p_role = clean_text(w.get("position_role", ""))
                dup = any(p_role and p_role == clean_text(e["position_role"]) for e in exist_works)
                if not dup and p_role:
                    cursor.execute("""
                    INSERT INTO work_histories (teacher_id, date_period, position_role)
                    VALUES (?, ?, ?)
                    """, (t_id, w.get("date_period", ""), p_role))
                    
            # Merge Teaching Assignments
            cursor.execute("SELECT * FROM teaching_assignments WHERE teacher_id = ?", (t_id,))
            exist_assigns = [dict(ta) for ta in cursor.fetchall()]
            for ta in rec.get("teaching_assignments", []):
                c_type = clean_text(ta.get("course_type", ""))
                s_name = clean_text(ta.get("subject_name", ""))
                s_cat = clean_text(ta.get("subject_category", ""))
                
                # Check duplicate
                dup = any(
                    c_type == clean_text(e["course_type"]) and s_name == clean_text(e["subject_name"])
                    for e in exist_assigns
                )
                if not dup and (c_type or s_name):
                    cursor.execute("""
                    INSERT INTO teaching_assignments (teacher_id, course_type, subject_category, subject_name)
                    VALUES (?, ?, ?, ?)
                    """, (t_id, c_type, s_cat, s_name))
                    
        else:
            # INSERT NEW RECORD
            new_count += 1
            cursor.execute("""
            INSERT INTO teachers (
                prefix_rank, first_name, last_name, age, position, 
                affiliation, workplace_address, phone, email, photo_url, notes, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            """, (
                rec.get("prefix_rank", ""),
                first,
                last,
                rec.get("age", ""),
                rec.get("position", ""),
                rec.get("affiliation", ""),
                rec.get("workplace_address", ""),
                rec.get("phone", ""),
                rec.get("email", ""),
                rec.get("photo_url", ""),
                rec.get("notes", "")
            ))
            t_id = cursor.lastrowid
            
            for edu in rec.get("educations", []):
                if any(str(v).strip() for v in edu.values()):
                    cursor.execute("""
                    INSERT INTO educations (teacher_id, year, level, degree_field, institution)
                    VALUES (?, ?, ?, ?, ?)
                    """, (t_id, edu.get("year", ""), edu.get("level", ""), edu.get("degree_field", ""), edu.get("institution", "")))
                    
            for tr in rec.get("trainings", []):
                if any(str(v).strip() for v in tr.values()):
                    cursor.execute("""
                    INSERT INTO trainings (teacher_id, year, course_name, organized_by)
                    VALUES (?, ?, ?, ?)
                    """, (t_id, tr.get("year", ""), tr.get("course_name", ""), tr.get("organized_by", "")))
                    
            for w in rec.get("work_histories", []):
                if any(str(v).strip() for v in w.values()):
                    cursor.execute("""
                    INSERT INTO work_histories (teacher_id, date_period, position_role)
                    VALUES (?, ?, ?)
                    """, (t_id, w.get("date_period", ""), w.get("position_role", "")))
                    
            for ta in rec.get("teaching_assignments", []):
                if any(str(v).strip() for v in ta.values()):
                    cursor.execute("""
                    INSERT INTO teaching_assignments (teacher_id, course_type, subject_category, subject_name)
                    VALUES (?, ?, ?, ?)
                    """, (t_id, ta.get("course_type", ""), ta.get("subject_category", ""), ta.get("subject_name", "")))
                    
    conn.commit()
    conn.close()
    
    return {
        "new_count": new_count,
        "merged_count": merged_count,
        "total_processed": len(teacher_records)
    }

def scan_directory(base_dir: str) -> Dict[str, Any]:
    all_records = []
    scanned_files = []
    errors = []
    
    print(f"[SCAN] Starting scan from: {base_dir}")
    
    for root, dirs, files in os.walk(base_dir):
        # Ignore system or venv folders
        if any(d in root for d in ['.venv', 'venv', 'uploads', '__pycache__', '.git', '.freebuff']):
            continue
        for f in files:
            path = os.path.join(root, f)
            if f.startswith('~$'): continue # Word/Excel temp lock files
            
            if f.lower().endswith('.docx'):
                scanned_files.append(path)
                try:
                    records = parse_docx_file(path)
                    print(f"[SCAN] Parsed {f}: {len(records)} records")
                    all_records.extend(records)
                except Exception as e:
                    error_msg = f"Error parsing {path}: {e}"
                    print(f"[SCAN] {error_msg}")
                    errors.append(error_msg)
                    
            elif f.lower().endswith('.xlsx'):
                scanned_files.append(path)
                try:
                    records = parse_xlsx_file(path)
                    print(f"[SCAN] Parsed {f}: {len(records)} records")
                    all_records.extend(records)
                except Exception as e:
                    error_msg = f"Error parsing {path}: {e}"
                    print(f"[SCAN] {error_msg}")
                    errors.append(error_msg)
    
    print(f"[SCAN] Total records found: {len(all_records)}")
    
    # Merge and save
    result = merge_and_save_data(all_records)
    result["scanned_files"] = scanned_files
    result["errors"] = errors
    return result
